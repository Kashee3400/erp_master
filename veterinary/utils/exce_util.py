from typing import Dict
from django.db import transaction
from tablib import Dataset
import pandas as pd
import logging
from openpyxl import load_workbook
from datetime import datetime, date
import re

from veterinary.resources.cattle_resources import CombinedCattleResource
from member.resources import UserDeviceResource,SahayakIncentivesResource,UserResource

RESOURCE_REGISTRY = {
    "veterinary.Animal": CombinedCattleResource,
    "auth.User":UserResource,
    "member.UserDevice":UserDeviceResource,
    "member.SahayakIncentives":SahayakIncentivesResource
}

logger = logging.getLogger(__name__)


class ExcelDataProcessor:
    """Utility class for processing Excel data"""

    @staticmethod
    def detect_data_types(df: pd.DataFrame) -> Dict[str, str]:
        """Detect data types for each column"""
        type_mapping = {}

        for column in df.columns:
            sample_data = df[column].dropna().head(10)

            if sample_data.empty:
                type_mapping[column] = 'text'
                continue

            # Check for dates
            elif ExcelDataProcessor._is_date_column(sample_data):
                type_mapping[column] = 'date'

            # Check for numbers
            elif ExcelDataProcessor._is_numeric_column(sample_data):
                type_mapping[column] = 'number'

            # Check for emails
            elif all('@' in str(val) and '.' in str(val) for val in sample_data):
                type_mapping[column] = 'email'

            # Default to text
            else:
                type_mapping[column] = 'text'

        return type_mapping

    @staticmethod
    def _is_date_column(sample_data) -> bool:
        """Check if column contains dates"""
        date_patterns = [
            r'\d{4}-\d{2}-\d{2}',
            r'\d{2}/\d{2}/\d{4}',
            r'\d{2}-\d{2}-\d{4}',
        ]

        for val in sample_data:
            val_str = str(val)
            if any(re.match(pattern, val_str) for pattern in date_patterns):
                return True
        return False

    @staticmethod
    def _is_numeric_column(sample_data) -> bool:
        """Check if column contains numbers"""
        try:
            pd.to_numeric(sample_data)
            return True
        except (ValueError, TypeError):
            return False

    @staticmethod
    def clean_and_validate_data(df: pd.DataFrame, field_mapping: Dict[str, str]) -> pd.DataFrame:
        """Clean and validate data based on field mapping"""
        cleaned_df = df.copy()

        for column, field_type in field_mapping.items():
            if column not in cleaned_df.columns:
                continue

            if field_type == 'boolean':
                cleaned_df[column] = cleaned_df[column].apply(
                    ExcelDataProcessor._convert_to_boolean
                )
            elif field_type == 'date':
                cleaned_df[column] = pd.to_datetime(
                    cleaned_df[column],
                    errors='coerce'
                ).dt.date
            elif field_type == 'number':
                cleaned_df[column] = pd.to_numeric(
                    cleaned_df[column],
                    errors='coerce'
                )
            elif field_type == 'email':
                cleaned_df[column] = cleaned_df[column].apply(
                    ExcelDataProcessor._validate_email
                )

        return cleaned_df

    @staticmethod
    def _convert_to_boolean(value):
        """Convert various representations to boolean"""
        if pd.isna(value):
            return None

        str_val = str(value).lower().strip()
        if str_val in ['true', 'yes', '1', 'y', 'on']:
            return True
        elif str_val in ['false', 'no', '0', 'n', 'off', '']:
            return False
        return None

    @staticmethod
    def _validate_email(value):
        """Basic email validation"""
        if pd.isna(value) or value == '':
            return None

        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
        if re.match(email_pattern, str(value)):
            return str(value)
        return None


def detect_column_types(resource_class) -> Dict:
    """Detect column types based on model fields and data patterns"""
    column_types = {}

    # Get model fields from resource
    if hasattr(resource_class._meta, 'model'):
        model = resource_class._meta.model
        for field in model._meta.fields:
            field_name = field.name

            # Map Django field types to pandas dtypes
            if field.__class__.__name__ in ['CharField', 'TextField', 'SlugField']:
                column_types[field_name] = 'text'
            elif field.__class__.__name__ in ['IntegerField', 'SmallIntegerField', 'BigIntegerField']:
                column_types[field_name] = 'integer'
            elif field.__class__.__name__ in ['FloatField', 'DecimalField']:
                column_types[field_name] = 'float'
            elif field.__class__.__name__ in ['DateField', 'DateTimeField']:
                column_types[field_name] = 'datetime'
            elif field.__class__.__name__ == 'BooleanField':
                column_types[field_name] = 'boolean'
            elif field.__class__.__name__ == 'ForeignKey':
                # Foreign keys often reference codes/IDs that should be preserved as text
                column_types[field_name] = 'text'

    return column_types


def process_import_enhanced(file_path, selected_sheets, target_model) -> Dict:
    """Enhanced import with proper type detection and preservation"""
    results = {}

    # Get resource class
    resource_class = RESOURCE_REGISTRY.get(target_model)
    if not resource_class:
        return {"error": f"No resource registered for {target_model}"}

    resource = resource_class()
    column_types = detect_column_types(resource_class)

    for sheet_name in selected_sheets:
        try:
            # Read with proper dtype detection
            df = read_excel_with_types(
                file_path,
                sheet_name,
                column_types,
            )

            # Remove empty rows
            df = df.dropna(how='all')

            if df.empty:
                results[sheet_name] = {
                    'created': 0,
                    'updated': 0,
                    'errors': ['Sheet is empty'],
                    'total_rows': 0,
                    'success': False
                }
                continue

            # Process with transaction for atomicity
            with transaction.atomic():
                result = _import_data_with_validation(
                    df, resource, column_types
                )
                results[sheet_name] = result

        except Exception as e:
            logger.error(f"Error processing sheet {sheet_name}: {e}", exc_info=True)
            results[sheet_name] = {
                'created': 0,
                'updated': 0,
                'errors': [f'Sheet processing failed: {str(e)}'],
                'total_rows': 0,
                'success': False
            }

    return results


def read_excel_with_types(file_path, sheet_name, column_types, pre_mapping=None) -> pd.DataFrame:
    """Read Excel with proper type preservation"""

    # First, use openpyxl to detect actual cell types
    wb = load_workbook(file_path, data_only=True, read_only=True)
    ws = wb[sheet_name]

    # Get headers
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    # Detect which columns contain text with leading zeros
    text_columns = set()
    sample_size = min(10, ws.max_row - 1)  # Sample first 10 data rows

    for row_idx in range(2, 2 + sample_size):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and isinstance(cell.value, str):
                # Check if it's a code-like value (digits with possible leading zeros)
                if re.match(r'^0\d+', str(cell.value)):
                    text_columns.add(header)

    wb.close()

    # Now read with pandas, specifying dtype for detected text columns
    dtype_spec = {}

    # Apply pre-mapping to column names if needed
    for header in headers:
        mapped_name = pre_mapping.get(header, header) if pre_mapping else header

        # Check if this column should be text
        if header in text_columns:
            dtype_spec[header] = str
        elif mapped_name in column_types:
            if column_types[mapped_name] == 'text':
                dtype_spec[header] = str
            elif column_types[mapped_name] == 'integer':
                dtype_spec[header] = 'Int64'  # Nullable integer
            elif column_types[mapped_name] == 'float':
                dtype_spec[header] = float

    # Read with specified dtypes
    df = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        dtype=dtype_spec,
        keep_default_na=True,
        na_values=['', 'NA', 'N/A', 'null', 'NULL']
    )

    # Additional processing for specific columns
    for col in df.columns:
        if col in dtype_spec and dtype_spec[col] == str:
            # Ensure strings are properly formatted
            df[col] = df[col].apply(lambda x: str(x).strip() if pd.notna(x) else None)

            # For code-like columns, ensure leading zeros are preserved
            if col in text_columns:
                df[col] = df[col].apply(
                    lambda x: x if x is None else str(x).zfill(len(str(x)))
                )

    return df


def _import_data_with_validation(df, resource, column_types):
    """Import data with proper validation and type handling"""

    # Convert DataFrame to Dataset
    dataset = Dataset()
    dataset.headers = df.columns.tolist()

    for idx, row in df.iterrows():
        row_data = []
        for col in df.columns:
            value = row[col]

            # Handle NaN/None
            if pd.isna(value):
                row_data.append(None)
                continue

            # Preserve string types explicitly
            if col in column_types and column_types[col] == 'text':
                # Ensure value is string and preserve formatting
                row_data.append(str(value))

            elif isinstance(value, (datetime, date)):
                # Always promote to datetime before passing forward
                if isinstance(value, date) and not isinstance(value, datetime):
                    value = datetime.combine(value, datetime.min.time())
                row_data.append(value)

            elif isinstance(value, (int, float)):
                # Check if this should be a string (e.g., codes)
                if col in column_types and column_types[col] == 'text':
                    row_data.append(
                        str(int(value)) if isinstance(value, float) and value.is_integer() else str(value))
                else:
                    row_data.append(value)
            else:
                row_data.append(str(value) if value is not None else None)

        dataset.append(row_data)

    # Perform dry run
    try:
        dry_result = resource.import_data(dataset, dry_run=True, raise_errors=False)
    except Exception as e:
        logger.error(f"Error importing dataset {dataset}: {e}", exc_info=True)

        return {
            'created': 0,
            'updated': 0,
            'errors': [f'Dry run failed: {str(e)}'],
            'total_rows': len(dataset),
            'success': False
        }

    # Collect validation errors
    if dry_result.has_validation_errors():
        validation_errors = []
        for idx, error in enumerate(dry_result.invalid_rows, start=1):
            validation_errors.append(f"Row {idx}: {error.error}")

        return {
            'created': 0,
            'updated': 0,
            'errors': validation_errors,
            'total_rows': len(dataset),
            'success': False
        }

    # Collect row errors
    if dry_result.has_errors():
        error_messages = []
        for idx, error in enumerate(dry_result.row_errors(), start=1):
            error_messages.append(f"Row {idx}: {error.error}")

        return {
            'created': 0,
            'updated': 0,
            'errors': error_messages,
            'total_rows': len(dataset),
            'success': False
        }

    # Perform actual import
    try:
        import_result = resource.import_data(dataset, dry_run=False, raise_errors=False)

        error_messages = []
        if import_result.has_errors():
            for error in import_result.row_errors():
                row_num = error[0] if isinstance(error, tuple) else error.row_number
                error_msg = error[1] if isinstance(error, tuple) else error.error
                error_messages.append(f"Row {row_num}: {error_msg}")

        # Count results by type
        created_count = 0
        updated_count = 0
        skipped_count = 0
        deleted_count = 0

        for row in import_result.rows:
            if hasattr(row, 'import_type'):
                if row.import_type == 'new':
                    created_count += 1
                elif row.import_type == 'update':
                    updated_count += 1
                elif row.import_type == 'skip':
                    skipped_count += 1
                elif row.import_type == 'delete':
                    deleted_count += 1

        return {
            'created': created_count,
            'updated': updated_count,
            'skipped': skipped_count,
            'deleted': deleted_count,
            'errors': error_messages,
            'total_rows': len(dataset),
            'success': not import_result.has_errors(),
            'totals': import_result.totals if hasattr(import_result, 'totals') else {}
        }

    except Exception as e:
        logger.error(f"Import execution failed: {e}", exc_info=True)
        return {
            'created': 0,
            'updated': 0,
            'errors': [f'Import failed: {str(e)}'],
            'total_rows': len(dataset),
            'success': False
        }
